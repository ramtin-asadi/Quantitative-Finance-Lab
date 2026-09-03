from __future__ import annotations

import hashlib
import io
import json
import re
import zipfile
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "data"
HERE = Path(__file__).resolve().parent
RAW = HERE / "raw"
CACHE = HERE / "cache"
API_CACHE = CACHE / "api"
STATE = CACHE / "state.json"
MARKET_OUTPUT = DATA / "finra_credit_market.parquet"
PRICING_OUTPUT = DATA / "finra_structured_pricing.parquet"
ACTIVITY_OUTPUT = DATA / "finra_structured_activity.parquet"
ERRATA_OUTPUT = DATA / "finra_structured_errata.parquet"
PARSER_VERSION = 5
DATE_PATTERN = re.compile(r"(20\d{6})")
METRIC_PATTERN = re.compile(
    r"price|trade|volume|size|spread|yield|coupon|balance|count|percentile|transaction|deviation",
    re.IGNORECASE,
)
SUBMETRIC_PATTERN = re.compile(r"customer|dealer|\$|^\s*[<>]=?", re.IGNORECASE)
FOOTER_PREFIXES = (
    "* indicates",
    "note:",
    "as of december",
    "effective june",
    "these reports",
    "for additional information",
)
PRICING_COLUMNS = [
    "report_date",
    "sheet",
    "table_title",
    "table_context",
    "row_group",
    "rating_group",
    "metric_group",
    "row_label",
    "column_label",
    "value",
    "value_status",
    "source_archive",
    "source_file",
]
ACTIVITY_COLUMNS = [
    "report_date",
    "sheet",
    "table_title",
    "row_group",
    "row_label",
    "column_label",
    "value",
    "value_status",
    "source_archive",
    "source_file",
]
ERRATA_CACHE_COLUMNS = [
    "report_date",
    "trade_date",
    "correction_date",
    "asset_class",
    "sub_asset_class",
    "note",
    "source_archive",
    "source_file",
]
MARKET_COLUMNS = [
    "date",
    "frequency",
    "market",
    "dataset",
    "is_144a",
    "security_category",
    "trade_side",
    "grade",
    "benchmark",
    "maturity_bucket",
    "source_field",
    "value",
]


def cell_text(value: object) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d")
    return " ".join(str(value).replace("\ufffd", "").replace("\u2020", "").strip().split())


def report_date(name: str) -> pd.Timestamp:
    match = DATE_PATTERN.search(name)
    if not match:
        raise ValueError(f"No YYYYMMDD report date in {name}")
    return pd.to_datetime(match.group(1), format="%Y%m%d")


def value_status(value: object) -> tuple[float | None, str | None]:
    if isinstance(value, (int, float)) and not isinstance(value, bool) and pd.notna(value):
        return float(value), "reported"
    text = cell_text(value)
    if text and set(text) <= {"*"}:
        return None, "suppressed"
    numeric = pd.to_numeric(text.replace(",", ""), errors="coerce")
    if pd.notna(numeric):
        return float(numeric), "reported"
    return None, None


def fill_right(values: list[str]) -> list[str]:
    current = ""
    filled = []
    for value in values:
        if value:
            current = value
        filled.append(current)
    return filled


def column_label(header_rows: list[list[str]], column: int) -> str:
    labels = []
    for row in header_rows:
        value = fill_right(row)[column]
        if not value or "metric" in value.lower() or value.lower() == "asset class":
            continue
        if not labels or labels[-1] != value:
            labels.append(value)
    return " | ".join(labels)


def numeric_cells(values: list[object], start: int) -> list[tuple[int, float | None, str]]:
    cells = []
    for column in range(start, len(values)):
        value, status = value_status(values[column])
        if status:
            cells.append((column, value, status))
    return cells


def next_label(frame: pd.DataFrame, row: int, label_column: int) -> str:
    for following in range(row + 1, min(row + 4, len(frame))):
        label = cell_text(frame.iat[following, label_column])
        if label:
            return label
    return ""


def footer_row(label: str) -> bool:
    return label.lower().startswith(FOOTER_PREFIXES)


def parse_pricing_sheet(
    frame: pd.DataFrame, sheet: str, archive_name: str, workbook_name: str
) -> list[dict[str, object]]:
    rows = []
    date = report_date(workbook_name)
    inside_table = False
    table_title = ""
    table_context: list[str] = []
    header_rows: list[list[str]] = []
    group_header: list[str] | None = None
    row_group = ""
    rating_group = ""
    metric_group = ""
    label_column = 1

    for row in range(len(frame)):
        raw_values = frame.iloc[row].tolist()
        values = [cell_text(value) for value in raw_values]
        titles = [value for value in values if "PRICING TABLE:" in value.upper()]
        if titles:
            label_column = next(index for index, value in enumerate(values) if value in titles)
            table_title = titles[0].split(":", 1)[-1].strip()
            table_context = []
            header_rows = []
            group_header = None
            row_group = ""
            rating_group = ""
            metric_group = ""
            inside_table = True
            continue
        if not inside_table:
            continue

        label = values[label_column]
        if footer_row(label):
            inside_table = False
            continue
        cells = numeric_cells(raw_values, label_column + 1)
        data_text = [value for value in values[label_column + 1 :] if value]
        generic_header = "metric" in label.lower()
        following = next_label(frame, row, label_column)
        group_row = bool(
            label
            and not generic_header
            and not METRIC_PATTERN.search(label)
            and not SUBMETRIC_PATTERN.search(label)
            and METRIC_PATTERN.search(following)
        )

        if group_row:
            row_group = label
            group_header = values
            continue
        if cells and label:
            upper = label.upper()
            if "VOLUME OF TRADES" in upper:
                metric_group = "volume"
            elif "NUMBER OF TRADES" in upper:
                metric_group = "count"
            elif METRIC_PATTERN.search(label):
                metric_group = ""
            active_headers = [*header_rows, *([group_header] if group_header else [])]
            for column, value, status in cells:
                label_for_column = column_label(active_headers, column)
                if not label_for_column:
                    continue
                rows.append(
                    {
                        "report_date": date,
                        "sheet": sheet,
                        "table_title": table_title,
                        "table_context": " | ".join(table_context),
                        "row_group": row_group,
                        "rating_group": rating_group,
                        "metric_group": metric_group,
                        "row_label": label,
                        "column_label": label_for_column,
                        "value": value,
                        "value_status": status,
                        "source_archive": archive_name,
                        "source_file": workbook_name,
                    }
                )
            continue
        if generic_header or (not label and data_text):
            section = label.split("/", 1)[0].strip().lower()
            if section in {"investment grade", "non-investment grade"}:
                rating_group = section.title()
            header_rows.append(values)
            continue
        if label and data_text:
            if not METRIC_PATTERN.search(label) and not SUBMETRIC_PATTERN.search(label):
                row_group = label
                group_header = values
            continue
        if label and not footer_row(label):
            table_context.append(label)

    return rows


def parse_activity_sheet(
    frame: pd.DataFrame, sheet: str, archive_name: str, workbook_name: str
) -> list[dict[str, object]]:
    rows = []
    date = report_date(workbook_name)
    label_column = 1
    header_rows: list[list[str]] = []
    row_group = ""
    for row in range(len(frame)):
        raw_values = frame.iloc[row].tolist()
        values = [cell_text(value) for value in raw_values]
        label = values[label_column]
        if footer_row(label):
            break
        cells = numeric_cells(raw_values, label_column + 1)
        data_text = [value for value in values[label_column + 1 :] if value]
        if label.upper() == "ASSET CLASS":
            header_rows = [values]
            row_group = ""
            continue
        if not label and data_text and not cells and header_rows:
            header_rows.append(values)
            continue
        if cells and label and header_rows:
            direct_asset_class = any(
                "INVESTMENT GRADE" in value.upper() for value in header_rows[0]
            ) and label.upper() in {"ABS", "CBO/CDO/CLO", "OTHER"}
            effective_group = "" if direct_asset_class else row_group
            for column, value, status in cells:
                label_for_column = column_label(header_rows, column)
                if not label_for_column:
                    continue
                rows.append(
                    {
                        "report_date": date,
                        "sheet": sheet,
                        "table_title": "Structured Product Trading Activity",
                        "row_group": effective_group,
                        "row_label": label,
                        "column_label": label_for_column,
                        "value": value,
                        "value_status": status,
                        "source_archive": archive_name,
                        "source_file": workbook_name,
                    }
                )
            continue
        if label and not cells:
            row_group = label
    return rows


def parse_errata(
    frame: pd.DataFrame, archive_name: str, workbook_name: str
) -> list[dict[str, object]]:
    header = None
    for row in range(len(frame)):
        text = [cell_text(value).lower() for value in frame.iloc[row].tolist()]
        if "trade date" in text and "date of correction" in text:
            header = row
            break
    if header is None:
        return []
    rows = []
    for row in range(header + 1, len(frame)):
        values = [cell_text(value) for value in frame.iloc[row].tolist()]
        nonempty = [value for value in values if value]
        if len(nonempty) < 4:
            continue
        trade_date = pd.to_datetime(nonempty[0], errors="coerce")
        correction_date = pd.to_datetime(nonempty[1], errors="coerce")
        if pd.isna(trade_date) or pd.isna(correction_date):
            continue
        rows.append(
            {
                "report_date": report_date(workbook_name),
                "trade_date": trade_date,
                "correction_date": correction_date,
                "asset_class": nonempty[2],
                "sub_asset_class": nonempty[3],
                "note": nonempty[4] if len(nonempty) > 4 else "",
                "source_archive": archive_name,
                "source_file": workbook_name,
            }
        )
    return rows


def read_workbook(content: bytes, name: str) -> dict[str, pd.DataFrame]:
    if name.lower().endswith(".xlsx"):
        book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        frames = {
            sheet: pd.DataFrame(book[sheet].iter_rows(values_only=True))
            for sheet in book.sheetnames
        }
        book.close()
        return frames
    excel = pd.ExcelFile(io.BytesIO(content), engine="xlrd")
    return {sheet: excel.parse(sheet_name=sheet, header=None) for sheet in excel.sheet_names}


def parse_archive(archive: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    pricing_rows = []
    activity_rows = []
    errata_rows = []
    with zipfile.ZipFile(archive) as bundle:
        for name in sorted(bundle.namelist()):
            if not name.lower().endswith((".xlsx", ".xls")):
                continue
            workbook_name = Path(name).name
            kind = "pricing" if "PXTABLES" in name.upper() else "activity" if "STAR" in name.upper() else None
            if kind is None:
                continue
            for sheet, frame in read_workbook(bundle.read(name), name).items():
                if sheet.lower() == "notes":
                    continue
                if sheet.lower() == "errata":
                    errata_rows.extend(parse_errata(frame, archive.name, workbook_name))
                elif kind == "pricing":
                    pricing_rows.extend(parse_pricing_sheet(frame, sheet, archive.name, workbook_name))
                else:
                    activity_rows.extend(parse_activity_sheet(frame, sheet, archive.name, workbook_name))

    pricing = pd.DataFrame(pricing_rows, columns=PRICING_COLUMNS).drop_duplicates()
    activity = pd.DataFrame(activity_rows, columns=ACTIVITY_COLUMNS).drop_duplicates()
    errata = pd.DataFrame(errata_rows, columns=ERRATA_CACHE_COLUMNS).drop_duplicates()
    pricing = pricing.sort_values(
        [
            "report_date",
            "sheet",
            "table_title",
            "table_context",
            "row_group",
            "rating_group",
            "row_label",
            "column_label",
        ],
        ignore_index=True,
    )
    activity = activity.sort_values(
        ["report_date", "row_group", "row_label", "column_label"], ignore_index=True
    )
    return pricing, activity, errata


def archive_digest(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def cache_paths(archive: Path) -> dict[str, Path]:
    stem = archive.stem
    return {
        "pricing": CACHE / f"{stem}_pricing.parquet",
        "activity": CACHE / f"{stem}_activity.parquet",
        "errata": CACHE / f"{stem}_errata.parquet",
    }


def write_cache(frame: pd.DataFrame, path: Path) -> None:
    temporary = path.with_suffix(".parquet.tmp")
    frame.to_parquet(temporary, index=False, compression="zstd")
    temporary.replace(path)


def snake_case(value: str) -> str:
    value = re.sub(r"(?<!^)(?=[A-Z])", "_", value).lower()
    return re.sub(r"[^a-z0-9]+", "_", value).strip("_")


def api_table(filename: str) -> pd.DataFrame:
    path = API_CACHE / filename
    if not path.exists():
        return pd.DataFrame()
    return pd.read_parquet(path)


def market_rows(
    frame: pd.DataFrame,
    measures: list[str],
    **dimensions: object,
) -> pd.DataFrame:
    if frame.empty:
        return pd.DataFrame(columns=MARKET_COLUMNS)
    base = pd.DataFrame(index=frame.index)
    for column in MARKET_COLUMNS[:-2]:
        default = pd.NA if column == "is_144a" else ""
        value = dimensions.get(column, default)
        base[column] = value
    pieces = []
    for measure in measures:
        if measure not in frame:
            continue
        piece = base.copy()
        piece["source_field"] = snake_case(measure)
        piece["value"] = pd.to_numeric(frame[measure], errors="coerce")
        pieces.append(piece.loc[piece["value"].notna()])
    if not pieces:
        return pd.DataFrame(columns=MARKET_COLUMNS)
    return pd.concat(pieces, ignore_index=True)[MARKET_COLUMNS]


def build_market_output() -> None:
    pieces = []
    breadth_specs = [
        ("finra_corporate_market_breadth.parquet", "corporate", False),
        ("finra_corporate_144a_market_breadth.parquet", "corporate", True),
        ("finra_agency_market_breadth.parquet", "agency", None),
    ]
    for filename, market, is_144a in breadth_specs:
        frame = api_table(filename)
        if frame.empty:
            continue
        pieces.append(
            market_rows(
                frame,
                [
                    "advances",
                    "declines",
                    "unchanged",
                    "totalVolume",
                    "totalTrades",
                    "fiftyTwoWeekHigh",
                    "fiftyTwoWeekLow",
                ],
                date=pd.to_datetime(frame["tradeReportDate"]),
                frequency="daily",
                market=market,
                dataset="breadth",
                is_144a=is_144a,
                security_category=frame["productCategory"].astype("string"),
            )
        )

    sentiment_specs = [
        ("finra_corporate_market_sentiment.parquet", "corporate", False),
        ("finra_corporate_144a_market_sentiment.parquet", "corporate", True),
        ("finra_agency_market_sentiment.parquet", "agency", None),
    ]
    for filename, market, is_144a in sentiment_specs:
        frame = api_table(filename)
        if frame.empty:
            continue
        pieces.append(
            market_rows(
                frame,
                ["totalVolume", "totalTransactions", "totalTrades"],
                date=pd.to_datetime(frame["tradeReportDate"]),
                frequency="daily",
                market=market,
                dataset="sentiment",
                is_144a=is_144a,
                security_category=frame["tradeType"].astype("string"),
                trade_side=frame["productCategory"].astype("string"),
            )
        )

    frame = api_table("finra_corporate_agency_capped_volume.parquet")
    if not frame.empty:
        frame["tradeReportDate"] = pd.to_datetime(frame["tradeReportDate"])
        frame = frame.sort_values("tradeReportDate").drop_duplicates(
            ["tradeYear", "tradeMonth", "gradeCode", "144AFlag"], keep="last"
        )
        identifiers = {"tradeReportDate", "tradeMonth", "tradeYear", "144AFlag", "gradeCode"}
        measures = [column for column in frame.columns if column not in identifiers]
        market = frame["gradeCode"].map(lambda value: "agency" if value == "AGCY" else "corporate")
        pieces.append(
            market_rows(
                frame,
                measures,
                date=pd.to_datetime(
                    {"year": frame["tradeYear"], "month": frame["tradeMonth"], "day": 1}
                ),
                frequency="monthly",
                market=market,
                dataset="capped_volume",
                is_144a=frame["144AFlag"].map({"Y": True, "N": False}).astype("boolean"),
                grade=frame["gradeCode"].astype("string"),
            )
        )

    treasury_specs = [
        ("finra_treasury_daily_aggregates.parquet", "tradeDate", "daily"),
        ("finra_treasury_monthly_aggregates.parquet", "beginningOfTheMonthDate", "monthly"),
    ]
    for filename, date_field, frequency in treasury_specs:
        frame = api_table(filename)
        if frame.empty:
            continue
        pieces.append(
            market_rows(
                frame,
                [
                    "dealerCustomerVolume",
                    "dealerCustomerCount",
                    "atsInterdealerVolume",
                    "atsInterdealerCount",
                    "volumeWeightedAveragePrice",
                ],
                date=pd.to_datetime(frame[date_field]),
                frequency=frequency,
                market="treasury",
                dataset="trading_activity",
                security_category=frame["productCategory"].astype("string"),
                benchmark=frame.get("benchmark", pd.Series("", index=frame.index)).fillna("").astype("string"),
                maturity_bucket=frame["yearsToMaturity"].fillna("").astype("string"),
            )
        )

    if not pieces:
        print("FINRA API cache is empty; finra_credit_market.parquet was not built")
        return
    result = pd.concat(pieces, ignore_index=True)
    for column in [
        "frequency",
        "market",
        "dataset",
        "security_category",
        "trade_side",
        "grade",
        "benchmark",
        "maturity_bucket",
        "source_field",
    ]:
        result[column] = result[column].fillna("").astype("string")
    result["is_144a"] = result["is_144a"].astype("boolean")
    result = result.drop_duplicates().sort_values(
        ["date", "market", "dataset", "security_category", "trade_side", "source_field"],
        ignore_index=True,
    )
    if result.empty or result["date"].isna().any() or result["source_field"].eq("").any():
        raise ValueError("Consolidated FINRA credit-market output failed validation")
    table = pa.Table.from_pandas(result, preserve_index=False).replace_schema_metadata(
        metadata("FINRA corporate-credit aggregates and fixed-income controls")
    )
    temporary = MARKET_OUTPUT.with_suffix(".parquet.tmp")
    pq.write_table(table, temporary, compression="zstd")
    temporary.replace(MARKET_OUTPUT)
    print(
        f"wrote {MARKET_OUTPUT} rows={len(result):,} "
        f"dates={result['date'].min().date()}..{result['date'].max().date()}"
    )


def metadata(dataset: str) -> dict[bytes, bytes]:
    values = {
        "dataset": dataset,
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "source": "FINRA historic reports and FINRA Developer Public API",
        "source_url": "https://www.finra.org/finra-data/browse-catalog/structured-product-activity-reports-and-tables/historic-reports",
        "api_url": "https://api.finra.org/data/group/fixedIncomeMarket/name/",
    }
    return {key.encode(): value.encode() for key, value in values.items()}


def latest_cache_date(paths: list[Path]) -> pd.Timestamp:
    for path in reversed(paths):
        dates = pd.read_parquet(path, columns=["report_date"])["report_date"]
        if not dates.empty:
            return pd.Timestamp(dates.max())
    raise RuntimeError("No report dates were found in FINRA archive caches")


def api_pricing_rows(after: pd.Timestamp) -> pd.DataFrame:
    frame = api_table("finra_cbo_cdo_clo_pricing.parquet")
    if frame.empty:
        return pd.DataFrame(columns=PRICING_COLUMNS)
    frame["reportDate"] = pd.to_datetime(frame["reportDate"])
    frame["recordLoadTimestamp"] = pd.to_datetime(frame["recordLoadTimestamp"], errors="coerce")
    key = [
        "reportDate",
        "pricingTableName",
        "productCode",
        "vintageCode",
        "categoryCode",
        "subCategoryCode",
    ]
    frame = frame.sort_values("recordLoadTimestamp").drop_duplicates(key, keep="last")
    frame = frame.loc[frame["reportDate"] > after].copy()
    if frame.empty:
        return pd.DataFrame(columns=PRICING_COLUMNS)
    report_text = frame["reportDate"].dt.strftime("%Y-%m-%d")
    column = frame["productCode"].fillna("").astype(str)
    vintage = frame["vintageCode"].fillna("").astype(str)
    column = column.where(vintage.eq("") | vintage.eq(report_text), column + " | " + vintage)
    result = pd.DataFrame(
        {
            "report_date": frame["reportDate"],
            "sheet": "CBO-CDO-CLO",
            "table_title": frame["pricingTableName"].fillna("CBO/CDO/CLO"),
            "table_context": "",
            "row_group": "",
            "rating_group": frame["productCode"].where(
                ~frame["productCode"].eq("CBO/CDO/CLO"), ""
            ),
            "metric_group": frame["categoryCode"].fillna(""),
            "row_label": frame["subCategoryCode"].fillna(""),
            "column_label": column,
            "value": pd.to_numeric(frame["valueAmount"], errors="coerce"),
            "value_status": frame["valueAmount"].notna().map(
                {True: "reported", False: "suppressed"}
            ),
            "source_archive": "FINRA_API",
            "source_file": "collateralizedObligationPricing",
        }
    )
    return result[PRICING_COLUMNS].drop_duplicates()


def activity_parts(product: object) -> tuple[str, str]:
    text = cell_text(product)
    match = re.fullmatch(r"(.+) \((P&I|IO/PO)\)", text)
    return (match.group(1), match.group(2)) if match else ("", text)


def api_activity_rows(after: pd.Timestamp) -> pd.DataFrame:
    pieces = []
    frame = api_table("finra_securitized_product_activity.parquet")
    if not frame.empty:
        frame["reportDate"] = pd.to_datetime(frame["reportDate"])
        frame["recordLoadTimestamp"] = pd.to_datetime(
            frame["recordLoadTimestamp"], errors="coerce"
        )
        key = ["reportDate", "productCode", "gradeCode", "categoryCode"]
        frame = frame.sort_values("recordLoadTimestamp").drop_duplicates(key, keep="last")
        frame = frame.loc[frame["reportDate"] > after].copy()
        groups_and_labels = frame["productCode"].map(activity_parts)
        category = frame["categoryCode"].replace(
            {
                "$ TRADES (000'S)": "$ TRADES | (000'S)",
                "TRADE COUNT": "TRADE | COUNT",
                "UNIQUE SEC ID'S": "UNIQUE | SEC ID'S",
            }
        )
        pieces.append(
            pd.DataFrame(
                {
                    "report_date": frame["reportDate"],
                    "sheet": "TradingActivity",
                    "table_title": "Structured Product Trading Activity",
                    "row_group": groups_and_labels.map(lambda value: value[0]),
                    "row_label": groups_and_labels.map(lambda value: value[1]),
                    "column_label": frame["gradeCode"].fillna("") + " | " + category.fillna(""),
                    "value": pd.to_numeric(frame["valueAmount"], errors="coerce"),
                    "value_status": frame["valueAmount"].notna().map(
                        {True: "reported", False: "suppressed"}
                    ),
                    "source_archive": "FINRA_API",
                    "source_file": "securitizedProductTradingActivity",
                }
            )
        )

    capped = api_table("finra_securitized_product_capped_volume.parquet")
    if not capped.empty:
        capped["tradeReportDate"] = pd.to_datetime(capped["tradeReportDate"])
        capped = capped.sort_values("tradeReportDate").drop_duplicates(
            ["tradeYear", "tradeMonth", "productType", "securitySubtype"], keep="last"
        )
        pieces.append(
            pd.DataFrame(
                {
                    "report_date": pd.to_datetime(
                        {
                            "year": capped["tradeYear"],
                            "month": capped["tradeMonth"],
                            "day": 1,
                        }
                    ),
                    "sheet": "CappedVolume",
                    "table_title": "Securitized Product Capped Volume",
                    "row_group": capped["productType"].fillna(""),
                    "row_label": capped["securitySubtype"].fillna(""),
                    "column_label": "AVERAGE TRANSACTION COUNT",
                    "value": pd.to_numeric(capped["averageTransactionCount"], errors="coerce"),
                    "value_status": capped["averageTransactionCount"].notna().map(
                        {True: "reported", False: "suppressed"}
                    ),
                    "source_archive": "FINRA_API",
                    "source_file": "securitizedProductsCappedVolume",
                }
            )
        )
    if not pieces:
        return pd.DataFrame(columns=ACTIVITY_COLUMNS)
    return pd.concat(pieces, ignore_index=True)[ACTIVITY_COLUMNS].drop_duplicates()


def api_errata_rows() -> pd.DataFrame:
    frame = api_table("finra_securitized_product_errata.parquet")
    if frame.empty:
        return pd.DataFrame(columns=ERRATA_CACHE_COLUMNS)
    result = pd.DataFrame(
        {
            "report_date": pd.to_datetime(frame["reportDate"]),
            "trade_date": pd.to_datetime(frame["tradeDate"]),
            "correction_date": pd.to_datetime(frame["correctionDate"]),
            "asset_class": frame["assetCode"].fillna(""),
            "sub_asset_class": frame["subAssetCode"].fillna(""),
            "note": frame["noteText"].fillna(""),
            "source_archive": "FINRA_API",
            "source_file": "securitizedProductErrata",
        }
    )
    return result[ERRATA_CACHE_COLUMNS].drop_duplicates()


def stream_caches(
    paths: list[Path], output: Path, dataset: str, extra: pd.DataFrame | None = None
) -> None:
    temporary = output.with_suffix(".parquet.tmp")
    writer = None
    schema = None
    rows = 0
    try:
        for path in paths:
            table = pq.read_table(path).replace_schema_metadata()
            if table.num_rows == 0:
                continue
            if writer is None:
                schema = table.schema
                writer = pq.ParquetWriter(
                    temporary, schema.with_metadata(metadata(dataset)), compression="zstd"
                )
            writer.write_table(table.cast(schema))
            rows += table.num_rows
        if writer is not None and extra is not None and not extra.empty:
            table = pa.Table.from_pandas(extra, schema=schema, preserve_index=False)
            writer.write_table(table)
            rows += table.num_rows
    finally:
        if writer is not None:
            writer.close()
    if writer is None:
        temporary.unlink(missing_ok=True)
        raise RuntimeError(f"No rows were available for {dataset}.")
    temporary.replace(output)
    print(f"wrote {output} rows={rows:,}")


def write_errata(paths: list[Path], extra: pd.DataFrame | None = None) -> None:
    frames = [pd.read_parquet(path) for path in paths]
    if extra is not None and not extra.empty:
        frames.append(extra)
    errata = pd.concat(frames, ignore_index=True)
    if errata.empty:
        ERRATA_OUTPUT.unlink(missing_ok=True)
        return
    event = ["trade_date", "correction_date", "asset_class", "sub_asset_class", "note"]
    compact = (
        errata.groupby(event, as_index=False, dropna=False)
        .agg(
            first_seen_report_date=("report_date", "min"),
            last_seen_report_date=("report_date", "max"),
            report_occurrences=("report_date", "size"),
        )
        .sort_values(event, ignore_index=True)
    )
    table = pa.Table.from_pandas(compact, preserve_index=False).replace_schema_metadata(
        metadata("FINRA structured-product data errata, unique correction events")
    )
    temporary = ERRATA_OUTPUT.with_suffix(".parquet.tmp")
    pq.write_table(table, temporary, compression="zstd")
    temporary.replace(ERRATA_OUTPUT)
    print(f"wrote {ERRATA_OUTPUT} rows={len(compact):,}")


def main() -> None:
    archives = sorted(RAW.glob("HISTORIC_SPREPORTS-*.zip"))
    if not archives:
        raise FileNotFoundError(
            f"No HISTORIC_SPREPORTS-YYYYMM.zip files under {RAW}. Download them from FINRA first."
        )
    CACHE.mkdir(parents=True, exist_ok=True)
    stored = json.loads(STATE.read_text()) if STATE.exists() else {}
    previous = stored.get("archives", {}) if stored.get("parser_version") == PARSER_VERSION else {}
    current = {}
    output_paths = {"pricing": [], "activity": [], "errata": []}

    for archive in archives:
        paths = cache_paths(archive)
        stat = archive.stat()
        old = previous.get(archive.name, {})
        unchanged = (
            old.get("size") == stat.st_size
            and old.get("mtime_ns") == stat.st_mtime_ns
            and all(path.exists() for path in paths.values())
        )
        if unchanged:
            digest = old["sha256"]
            print(f"cached {archive.name}")
        else:
            digest = archive_digest(archive)
            if old.get("sha256") == digest and all(path.exists() for path in paths.values()):
                print(f"cached {archive.name} (content unchanged)")
            else:
                pricing, activity, errata = parse_archive(archive)
                if pricing.empty or activity.empty:
                    raise RuntimeError(f"{archive.name} produced no pricing or activity rows.")
                write_cache(pricing, paths["pricing"])
                write_cache(activity, paths["activity"])
                write_cache(errata, paths["errata"])
                print(
                    f"parsed {archive.name}: pricing={len(pricing):,} "
                    f"activity={len(activity):,} errata={len(errata):,}"
                )
        current[archive.name] = {
            "sha256": digest,
            "size": stat.st_size,
            "mtime_ns": stat.st_mtime_ns,
        }
        for kind, path in paths.items():
            output_paths[kind].append(path)
        STATE.write_text(
            json.dumps({"parser_version": PARSER_VERSION, "archives": current}, indent=2, sort_keys=True)
            + "\n"
        )

    pricing_cutoff = latest_cache_date(output_paths["pricing"])
    activity_cutoff = latest_cache_date(output_paths["activity"])
    stream_caches(
        output_paths["pricing"],
        PRICING_OUTPUT,
        "FINRA structured-product pricing tables",
        api_pricing_rows(pricing_cutoff),
    )
    stream_caches(
        output_paths["activity"],
        ACTIVITY_OUTPUT,
        "FINRA structured-product trading activity",
        api_activity_rows(activity_cutoff),
    )
    write_errata(output_paths["errata"], api_errata_rows())
    build_market_output()


if __name__ == "__main__":
    main()
